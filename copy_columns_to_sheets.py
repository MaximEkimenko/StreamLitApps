import openpyxl


def copy_columns_to_sheets(file_path: str, source_sheet_name: str, exclude_sheets: tuple = None):
    """
    Копирование колонок из одного листа source_sheet_name excel файла file_path за исключением списка
    exclude_sheets листов
    """
    # TODO вынести имена колонок для копирования во внешние параметры
    # Открываем файл с помощью openpyxl
    book = openpyxl.load_workbook(file_path)
    source_sheet = book[source_sheet_name]
    result_str = []
    # Определение листов, которые необходимо исключить
    if exclude_sheets is None:
        exclude_sheets = tuple()

    # Считывание данных из исходного листа
    source_data = {}
    for row in source_sheet.iter_rows(min_row=4, values_only=True):
        p_num = row[0]
        # TODO подставить новые координаты значений для копирования
        next_sz = row[1]  # ссылка на след СЗ
        tech_queue = row[2]  # очерёдность
        rc_number = row[5]  # номер рабочего центра
        rc_analog_1 = row[6]  # аналог рабочего центра №1
        rc_analog_2 = row[7]  # аналог рабочего центра №2
        rc_analog_3 = row[8]  # аналог рабочего центра №3
        terminal_number = row[9]  # номер терминала

        if p_num is not None:
            source_data[p_num] = (tech_queue,
                                  next_sz,
                                  rc_number,
                                  rc_analog_1,
                                  rc_analog_2,
                                  rc_analog_3,
                                  terminal_number,
                                  )

    # Проход по каждому листу
    for sheet_name in book.sheetnames:
        if sheet_name != source_sheet_name and sheet_name not in exclude_sheets:
            sheet = book[sheet_name]
            # Запись данных в соответствующие ячейки
            # TODO откорректировать интервал
            for row_idx, row in enumerate(sheet.iter_rows(min_row=4, max_col=1, values_only=True), start=4):
                p_num = row[0]
                if p_num in source_data:
                    (tech_queue,
                     next_sz,
                     rc_number,
                     rc_analog_1,
                     rc_analog_2,
                     rc_analog_3,
                     terminal_number,
                     ) = source_data[p_num]
                    # Пропускаем объединенные ячейки
                    for cell in sheet[row_idx]:
                        if cell.coordinate in sheet.merged_cells:
                            continue
                        # TODO подставить реальные номера колонок
                        elif cell.column == 2:  # Следующее СЗ
                            cell.value = next_sz
                        if cell.column == 3:  # Очерёдность
                            cell.value = tech_queue
                        elif cell.column == 6:  # Номер РЦ
                            cell.value = rc_number
                        elif cell.column == 7:  # аналог РЦ 1
                            cell.value = rc_analog_1
                        elif cell.column == 8:  # аналог РЦ 2
                            cell.value = rc_analog_2
                        elif cell.column == 9:  # аналог РЦ 3
                            cell.value = rc_analog_3
                        elif cell.column == 10:  # аналог РЦ 3
                            cell.value = terminal_number
                    if tech_queue:
                        # строка успехов для вывода в streamlit
                        success_copy_str = (f"Скопировано в {sheet_name}: {p_num} - Очередность = {tech_queue},"
                                            f" Следующее СЗ = {next_sz}"
                                            f" Рабочие центры со своими аналогами  = "
                                            f"{rc_number, rc_analog_1, rc_analog_2, rc_analog_3}."
                                            f" Номер терминала = {terminal_number}.")
                        result_str.append(success_copy_str)
    book.save(file_path)
    book.close()
    return result_str


if __name__ == '__main__':
    # tst_file_path = r'C:\Users\user-18\Desktop\1\Трудоемкость серия М (в работе).xlsx'
    # tst_source_sheet_name = '12000М+'
    # tst_exclude_sheets = ('Интерполяция М', 'Интерполяция R', 'Лист1', 'Sheet1', 'Интерполяция SV', 'Интерполяция P')
    # copy_columns_to_sheets(tst_file_path, tst_source_sheet_name, tst_exclude_sheets)
    pass
